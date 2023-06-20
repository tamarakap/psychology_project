import os
import numpy as np
import torchvision
from tqdm import tqdm
from PIL import Image
import matplotlib.pyplot as plt
import clip
from argparse import ArgumentParser
from openpyxl import load_workbook
import torch
from torchvision import transforms
from training.dataset_utils import Rescale
from facenet_pytorch import MTCNN


my_path = os.path.abspath(os.path.dirname(__file__))

# output paths
context_output_path_clip = os.path.join(my_path, "..\outputs\context_out_sheet_clip.xlsx")
context_output_path_vgg = os.path.join(my_path, "..\outputs\context_out_sheet_vgg.xlsx")
thatcher_output_path_clip = os.path.join(my_path, "..\outputs\\thatcher_out_sheet_clip.xlsx")
thatcher_output_path_vgg = os.path.join(my_path, "..\outputs\\thatcher_out_sheet_vgg.xlsx")
occlusions_output_path_clip = os.path.join(my_path, "..\outputs\occlusions_out_sheet_clip.xlsx")
occlusions_output_path_vgg = os.path.join(my_path, "..\outputs\occlusions_out_sheet_vgg.xlsx")

#folders paths
context_folder_path = os.path.join(my_path, "..\celebrities_for_context_test")
thatcher_folder_path = os.path.join(my_path, "..\img_by_identity")
occlusions_folder_path = os.path.join(my_path, "..\occlusions")

weights_path = os.path.join(my_path, "..\\vgg_weights\\9.pth")

mtcnn = MTCNN(image_size=224)
normalize = transforms.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225])
composed_transforms = transforms.Compose([Rescale((160, 160)), normalize])
cos = torch.nn.CosineSimilarity()


def verify_rgb(img_path, img):
    if img.mode != "RGB":
        rgb_image = Image.new("RGB", img.size, (255, 255, 255))
        rgb_image.paste(img, mask=img.split()[3])
        rgb_image.save(img_path, "JPEG")
        img = Image.open(img_path)
    return img


def get_vgg_model():
    model = torchvision.models.vgg16().eval()
    model.features = torch.nn.DataParallel(model.features)
    model.classifier[-1] = torch.nn.Linear(in_features=4096, out_features=20)
    weights = torch.load(weights_path, map_location=torch.device('cpu'))['state_dict']
    model.load_state_dict(weights)

    return model


def get_clip_model():
    model, preprocess = clip.load("ViT-B/32", args.device)
    return model, preprocess


def get_model(network_type):
    if network_type == "clip":
        return get_clip_model()
    else:
        return get_vgg_model(), None


def calculate_thacher_index(dif_up, dif_inv):
    return (dif_up - dif_inv) / (dif_up + dif_inv)


def get_image_embedding(img_path, model, rotate=False, network_type="vgg", preprocess=None):
    img = Image.open(img_path)

    if network_type == "vgg":
        img = verify_rgb(img_path=img_path, img=img)
        img = mtcnn(img)

        if rotate:
            img = img.permute(1, 2, 0)  # change tensor shape to HxWxC
            img = torch.rot90(img, 2)
            img = img.permute(2, 0, 1)
        return model(img.unsqueeze(0).float())



    elif network_type == "clip":
        if rotate:
            img = img.rotate(180)
        image = preprocess(img).unsqueeze(0).to(args.device)

        with torch.no_grad():
            image_features = model.encode_image(image)

        plt.imshow(img)
        plt.axis('off')
        plt.show()

        return image_features


def load_data_for_thacher(data_dir):
    data_paths_list_up = []
    data_paths_list_inv = []
    data_dir = os.path.join(thatcher_folder_path, data_dir)
    imgs = os.listdir(data_dir)

    for img in imgs[:2]:
        im_path = os.path.join(data_dir, img)
        data_paths_list_up.append(im_path)
    for img in imgs[2:]:
        im_path = os.path.join(data_dir, img)
        data_paths_list_inv.append(im_path)

    return data_paths_list_up, data_paths_list_inv


def load_data_for_context_test():
    return os.listdir(context_folder_path)


def load_data_for_occlusions_test():
    return os.listdir(occlusions_folder_path)


def get_dataset_path(dataset_paths_list):
    if args.test_type == "context":
        return os.path.join(context_folder_path, dataset_paths_list)
    else:
        return os.path.join(occlusions_folder_path, dataset_paths_list)


def get_occlusions_test_results(dataset_paths_list, model, network_type, preprocess):
    rdm = np.zeros(2)
    dataset_paths_list = get_dataset_path(dataset_paths_list)
    imgs = os.listdir(dataset_paths_list)

    bottom_cover = os.path.join(dataset_paths_list, imgs[0])
    top_cover = os.path.join(dataset_paths_list, imgs[1])
    regular = os.path.join(dataset_paths_list, imgs[2])

    regular_img_embedding = get_image_embedding(regular, model, network_type=network_type,
                                                preprocess=preprocess)
    top_cover_img_embedding = get_image_embedding(top_cover, model, network_type=network_type, preprocess=preprocess)
    bottom_cover_img_embedding = get_image_embedding(bottom_cover, model, network_type=network_type,
                                                     preprocess=preprocess)
    rdm[0] = torch.cdist(regular_img_embedding, top_cover_img_embedding, 2)
    rdm[1] = torch.cdist(regular_img_embedding, bottom_cover_img_embedding, 2)
    return rdm


def get_context_test_results(dataset_paths_list, model, network_type, preprocess):
    rdm = np.zeros(2)
    dataset_paths_list = get_dataset_path(dataset_paths_list)
    imgs = os.listdir(dataset_paths_list)

    regular_img = os.path.join(dataset_paths_list, imgs[0])
    conext_img = os.path.join(dataset_paths_list, imgs[1])
    out_of_context_img = os.path.join(dataset_paths_list, imgs[2])

    regular_img_embedding = get_image_embedding(regular_img, model, network_type=network_type,
                                                preprocess=preprocess)
    conext_img_embedding = get_image_embedding(conext_img, model, network_type=network_type, preprocess=preprocess)
    out_of_context_img_embedding = get_image_embedding(out_of_context_img, model, network_type=network_type,
                                                       preprocess=preprocess)
    rdm[0] = torch.cdist(regular_img_embedding, conext_img_embedding, 2)
    rdm[1] = torch.cdist(regular_img_embedding, out_of_context_img_embedding, 2)
    return rdm


def get_thacher_results(dataset_paths_list_up, data_paths_list_inv, model, network_type, preprocess):
    results = np.zeros(3)

    img_path1 = dataset_paths_list_up[0]
    img_path2 = dataset_paths_list_up[1]

    img_embedding1 = get_image_embedding(img_path1, model, network_type=network_type, preprocess=preprocess)
    img_embedding2 = get_image_embedding(img_path2, model, network_type=network_type, preprocess=preprocess)

    # verify cdist
    results[0] = torch.cdist(img_embedding1, img_embedding2, 2,
                             compute_mode='use_mm_for_euclid_dist_if_necessary')

    img_path_inv1 = data_paths_list_inv[0]
    img_path_inv2 = data_paths_list_inv[1]
    img_embedding1 = get_image_embedding(img_path_inv1, model, rotate=True, network_type=network_type,
                                         preprocess=preprocess)
    img_embedding2 = get_image_embedding(img_path_inv2, model, rotate=True, network_type=network_type,
                                         preprocess=preprocess)

    results[1] = torch.cdist(img_embedding1, img_embedding2, 2)
    results[2] = calculate_thacher_index(results[0], results[1])

    return results


def thatcher_test(folder_path, model, network_type, preprocess):
    row_indx = 2
    thatcher_path = thatcher_output_path_vgg if network_type == "vgg" else thatcher_output_path_clip
    out = load_workbook(thatcher_path)
    thatcher_results_sheet = out.active
    for i, class_path in tqdm(enumerate(folder_path)):
        data_paths_list_up, data_paths_list_inv = load_data_for_thacher(class_path)
        results = get_thacher_results(data_paths_list_up, data_paths_list_inv, model, network_type, preprocess)
        thatcher_results_sheet.cell(row=row_indx, column=1).value = data_paths_list_up[0].split('\\')[-1]
        thatcher_results_sheet.cell(row=row_indx, column=2).value = data_paths_list_up[1].split('\\')[-1]
        thatcher_results_sheet.cell(row=row_indx, column=3).value = data_paths_list_inv[0].split('\\')[-1]
        thatcher_results_sheet.cell(row=row_indx, column=4).value = data_paths_list_inv[1].split('\\')[-1]
        thatcher_results_sheet.cell(row=row_indx, column=5).value = results[0]
        thatcher_results_sheet.cell(row=row_indx, column=6).value = results[1]
        thatcher_results_sheet.cell(row=row_indx, column=7).value = results[2]
        row_indx += 1
    out.save(thatcher_path)


def context_test(folder_path, model, network_type, preprocess):
    row_indx = 2
    context_path = context_output_path_vgg if network_type == "vgg" else context_output_path_clip
    out = load_workbook(context_path)
    context_results_sheet = out.active
    for i, class_path in tqdm(enumerate(folder_path)):
        context_results = get_context_test_results(class_path, model, network_type, preprocess)
        context_results_sheet.cell(row=row_indx, column=1).value = context_results[0]
        context_results_sheet.cell(row=row_indx, column=2).value = context_results[1]
        row_indx += 1
    out.save(context_path)


def occlusions_test(folder_path, model, network_type, preprocess):
    row_indx = 2
    occlusions_path = occlusions_output_path_vgg if network_type == "vgg" else occlusions_output_path_clip
    out = load_workbook(occlusions_path)

    context_results_sheet = out.active
    for i, class_path in tqdm(enumerate(folder_path)):
        occlusions_results = get_occlusions_test_results(class_path, model, network_type, preprocess)
        context_results_sheet.cell(row=row_indx, column=1).value = occlusions_results[0]
        context_results_sheet.cell(row=row_indx, column=2).value = occlusions_results[1]
        row_indx += 1
    out.save(occlusions_path)


# just paste this as running configuration:
# -test_type
# thatcher/context/occlusions
if __name__ == '__main__':
    parser = ArgumentParser()
    parser.add_argument("-test_type", "--test_type", dest="test_type", help="The type of test to perform")
    parser.add_argument("-network_type", "--network_type", dest="network_type",
                        help="the neural network that will be used: vgg or clip")
    parser.add_argument("-device", default="cpu")
    args = parser.parse_args()
    model, preprocess = get_model(args.network_type)

    # resize images!
    if args.test_type == "thatcher":
        thatcher_test(os.listdir(thatcher_folder_path), model, args.network_type, preprocess)

    elif args.test_type == "context":
        data_paths_list = load_data_for_context_test()
        context_test(data_paths_list, model, args.network_type, preprocess)

    elif args.test_type == "occlusions":
        data_paths_list = load_data_for_occlusions_test()

        # the occlusion test is exactly the same as the context test!
        occlusions_test(data_paths_list, model, args.network_type, preprocess)
    else:
        raise Exception("invalid test type!")
